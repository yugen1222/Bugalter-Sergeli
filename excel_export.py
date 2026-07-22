\
from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference


def create_month_excel(rows: list[dict], year: int, month: int) -> Path:
    output = Path(__file__).with_name(f"material_report_{year}_{month:02d}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Детали"

    headers = [
        "Дата", "Сотрудник", "Смена", "Выравнивание",
        "День закрыт", "Товар", "Количество", "Причина",
        "Ответственный", "Комментарий", "Статус"
    ]
    ws.append(headers)

    for row in rows:
        ws.append([
            row.get("check_date"),
            row.get("employee_name"),
            row.get("shift") or "",
            "Да" if row.get("alignment_done") else "Нет",
            "Да" if row.get("completed") else "Нет",
            row.get("item_name") or "",
            row.get("quantity") or 0,
            row.get("reason") or "",
            row.get("responsible_name") or "",
            row.get("comment") or "",
            "Найдено" if row.get("status") == "resolved" else (
                "Не найдено" if row.get("status") == "unresolved" else ""
            ),
        ])

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    widths = [13, 22, 12, 16, 14, 24, 12, 26, 22, 40, 15]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    summary = wb.create_sheet("Аналитика")
    summary["A1"] = f"Аналитика за {month:02d}.{year}"
    summary["A1"].font = Font(size=16, bold=True)

    product_stats = {}
    reason_stats = {}
    shift_stats = {}

    for row in rows:
        if row.get("item_name"):
            product_stats[row["item_name"]] = product_stats.get(row["item_name"], 0) + float(row.get("quantity") or 0)
        if row.get("reason"):
            reason_stats[row["reason"]] = reason_stats.get(row["reason"], 0) + 1
        shift = row.get("shift") or "Не указана"
        if row.get("item_name"):
            shift_stats[shift] = shift_stats.get(shift, 0) + 1

    summary.append([])
    summary.append(["Товар", "Общий минус"])
    for name, qty in sorted(product_stats.items(), key=lambda x: x[1], reverse=True):
        summary.append([name, qty])

    reason_start = summary.max_row + 3
    summary.cell(reason_start, 1, "Причина")
    summary.cell(reason_start, 2, "Количество случаев")
    for name, count in sorted(reason_stats.items(), key=lambda x: x[1], reverse=True):
        summary.append([name, count])

    shift_start = summary.max_row + 3
    summary.cell(shift_start, 1, "Смена")
    summary.cell(shift_start, 2, "Количество случаев")
    for name, count in sorted(shift_stats.items(), key=lambda x: x[1], reverse=True):
        summary.append([name, count])

    for row_idx in [3, reason_start, shift_start]:
        for cell in summary[row_idx]:
            if cell.column <= 2:
                cell.fill = header_fill
                cell.font = header_font

    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 20

    if product_stats:
        chart = BarChart()
        chart.title = "Товары с наибольшими минусами"
        data = Reference(summary, min_col=2, min_row=3, max_row=2 + len(product_stats))
        cats = Reference(summary, min_col=1, min_row=4, max_row=2 + len(product_stats))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 14
        summary.add_chart(chart, "D3")

    if reason_stats:
        pie = PieChart()
        pie.title = "Причины минусов"
        data = Reference(summary, min_col=2, min_row=reason_start, max_row=reason_start + len(reason_stats))
        labels = Reference(summary, min_col=1, min_row=reason_start + 1, max_row=reason_start + len(reason_stats))
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.height = 8
        pie.width = 12
        summary.add_chart(pie, "D20")

    wb.save(output)
    return output
